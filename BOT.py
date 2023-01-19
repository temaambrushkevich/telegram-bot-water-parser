import time
import shutil
import re
from threading import Thread

import telebot
from telebot import types
import TOKEN

from openpyxl import load_workbook

from docx2pdf import convert

from PARSING import parsing_price, parsing_stock
import schedule

bot = telebot.TeleBot(TOKEN.TOKEN)

@bot.message_handler(commands=['start'])
def start(message):
    bot.send_message(message.chat.id, "Приветствую, {0.first_name}!".format(message.from_user, bot.get_me()),
                     parse_mode='html')
    call_buttons_1(message)

# обработать нажатия на кнопки
@bot.message_handler(content_types=['text'])
def use_buttons(message):
    if message.chat.type == 'private':
        if message.text == "Цены":
            call_buttons_2(message)
        elif message.text == "Цены за определенный месяц":
            botmsg = "Введите месяц и год в формате \"месяц.год\", например 12.2022"
            a = telebot.types.ReplyKeyboardRemove()
            bot.send_message(message.chat.id, botmsg, reply_markup=a)

            bot.register_next_step_handler(message, get_month_prices_data)
        elif message.text == "Цены за все время":
            try:
                file = open('цены за всё время.xlsx', 'rb')
                bot.send_document(message.chat.id, file)
                call_buttons_1(message)
            except:
                print()
                error(message)

        if message.text == "Назад":
            call_buttons_1(message)

        if message.text == "Акции":
            call_buttons_3(message)
        elif message.text == "Акции за определенный месяц":
            botmsg = "Введите месяц и год в формате \"месяц.год\", например 12.2022"
            a = telebot.types.ReplyKeyboardRemove()
            bot.send_message(message.chat.id, botmsg, reply_markup=a)
            bot.register_next_step_handler(message, get_month_stocks_data)
        elif message.text == "Акции за все время":
            a = telebot.types.ReplyKeyboardRemove()
            bot.send_message(message.chat.id, "Ожидайте...", reply_markup=a)
            convert('все акции.docx')
            file1 = open('все акции.docx', 'rb')
            bot.send_document(message.chat.id, file1)
            file2 = open('все акции.pdf', 'rb')
            bot.send_document(message.chat.id, file2)
            call_buttons_1(message)

# получить данные за определенный месяц
def get_month_prices_data(message):
    def format_xlfile(date):
        file_name = "цены за всё время.xlsx"
        months = ["Январь", "Февраль", "Март", "Апрель", "Май", "Июнь", "Июль", "Август", "Сентябрь", "Октябрь",
                  "Ноябрь",
                  "Декабрь"]

        m_and_y = date
        m = date[0:2]
        y = date[3:7]
        print(m_and_y)
        print("Сегодня: " + date + '\n')

        month_file_name = "цены " + months[int(m) - 1] + " " + y + ".xlsx"

        # форматируем отдельные листы
        def format_separate_sheets():
            # копируем файл "цены.xlsx" в новый файл с названием "цены месяц.год.xlsx"
            shutil.copyfile(file_name, month_file_name)
            # открываем новый файл
            xlfile = load_workbook(month_file_name)

            # ищем листы с таким месяцем и годом
            i = 0
            for sheet_name in xlfile.sheetnames:
                if i != 0:
                    # if sheet_name[3:10] != date[3:10]:
                    if sheet_name[3:10] != m_and_y:
                        # удаляем ненужный лист
                        del xlfile[sheet_name]
                        xlfile.save(month_file_name)
                i += 1
        # форматируем общий лист
        def format_common_sheet():
            xlfile = load_workbook(month_file_name)
            sheet = xlfile["Общая от 2-х"]
            del_list = []
            i = 2
            while sheet.cell(row=i, column=1).value != None:
                # если дата совпадает с датой строки
                if m_and_y == (sheet.cell(row=i, column=1).value)[3:10]:
                    print(sheet.cell(row=i, column=1).value, "\t", (sheet.cell(row=i, column=1).value)[3:10])
                else:
                    del_list.append(i)
                i += 1
            # sheet.delete_rows()
            xlfile.save(month_file_name)

            # Удаляем лишние строки, не относящиеся к выбранному месяцу
            print("строки на удаление: ", del_list)
            for i in reversed(del_list):
                sheet.delete_rows(i)
            xlfile.save(month_file_name)
        # форматируем первый лист с датой, чтобы поправить ссылку
        def format_one_list():
            xlfile = load_workbook(month_file_name)
            sheet_name = xlfile.sheetnames[1]
            sheet = xlfile[sheet_name]
            i = 2
            while sheet.cell(row=4, column=i).value != None:
                sheet.cell(row=4, column=i).value = ''
                i += 1
            xlfile.save(month_file_name)


        format_separate_sheets()
        format_common_sheet()
        format_one_list()

        return month_file_name
    def check_date(m_and_y):
        all_file_name = "цены за всё время.xlsx"
        xlfile = load_workbook(all_file_name)
        sheet_name = xlfile.sheetnames[0]
        sheet = xlfile[sheet_name]

        i = 2
        while sheet.cell(row=i, column=1).value != None:
            # если дата совпадает с датой строки
            if m_and_y == str((sheet.cell(row=i, column=1).value)[3:10]):
                return "OK"
            i += 1
        return "NO"
    a = telebot.types.ReplyKeyboardRemove()
    # верный ввод
    if re.search(r"(([0]{1}\d{1})|([1][0-2]))[.][2][0][\d][\d]", message.text):
        bot.send_message(message.chat.id, "Ожидайте...", reply_markup=a)
        if check_date(message.text) == "OK":
            file = open(format_xlfile(message.text), 'rb')
            bot.send_document(message.chat.id, file)
        else:
            bot.send_message(message.chat.id, "Измерений за этот период не найдено!", reply_markup=a)
    # неверный ввод
    else:
        bot.send_message(message.chat.id, "Неверный ввод!", reply_markup=a)
    call_buttons_1(message)

def get_month_stocks_data(message):
    def check_date_in_bd(date):
        # проверяем дату в файле db.txt
        bdfile = open("bd.txt", "r+")
        if date + '\n' in bdfile:
            print("дата найдена в bd.txt")
            return "OK"
        else:
            print("Дата не найдена в bd.txt")
            return "NO"

    months = ["Январь", "Февраль", "Март", "Апрель", "Май", "Июнь", "Июль", "Август", "Сентябрь", "Октябрь", "Ноябрь",
              "Декабрь"]

    a = telebot.types.ReplyKeyboardRemove()
    if re.search(r"(([0]{1}\d{1})|([1][0-2]))[.][2][0][\d][\d]", message.text):
        bot.send_message(message.chat.id, "Ожидайте...", reply_markup=a)
        if check_date_in_bd(message.text) == "OK":
            filename1 = "акции " + months[int(message.text[0:2])-1] + " " + message.text[3:7] + ".docx"
            convert(filename1)
            file1 = open(filename1, 'rb')
            bot.send_document(message.chat.id, file1)

            filename2 = "акции " + months[int(message.text[0:2]) - 1] + " " + message.text[3:7] + ".pdf"
            file2 = open(filename2, 'rb')
            bot.send_document(message.chat.id, file2)
        else:
            bot.send_message(message.chat.id, "Акций за этот период не найдено!", reply_markup=a)
    call_buttons_1(message)

# вызвать кнопки
@bot.message_handler(content_types=['text'])
def call_buttons_1(message):
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)

    item1 = types.KeyboardButton("Цены")
    item2 = types.KeyboardButton("Акции")

    markup.add(item1, item2)

    bot.send_message(message.chat.id, "Выберите кнопку".format(message.from_user, bot.get_me()),
                     parse_mode='html', reply_markup=markup)
@bot.message_handler(content_types=['text'])
# кнопки меню - цены
def call_buttons_2(message):
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    item2 = types.KeyboardButton("Цены за определенный месяц")
    item3 = types.KeyboardButton("Цены за все время")
    item4 = types.KeyboardButton("Назад")
    markup.add(item2, item3, item4)
    bot.send_message(message.chat.id, "Выберите кнопку".format(message.from_user, bot.get_me()),
                     parse_mode='html', reply_markup=markup)
# кнопки меню - акции
def call_buttons_3(message):
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    item2 = types.KeyboardButton("Акции за определенный месяц")
    item3 = types.KeyboardButton("Акции за все время")
    item4 = types.KeyboardButton("Назад")
    markup.add(item2, item3, item4)
    bot.send_message(message.chat.id, "Выберите кнопку".format(message.from_user, bot.get_me()),
                     parse_mode='html', reply_markup=markup)

# Функция ошибок парсинга
@bot.message_handler(content_types=['text'])
def error_msg_parsing(message, text_msg):
    bot.send_message(message.chat.id,
                     text_msg.format(message.from_user, bot.get_me()),
                     parse_mode='html')

@bot.message_handler(content_types=['text'])
def error(message):
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    item1 = types.KeyboardButton("/start")

    markup.add(item1)

    bot.send_message(message.chat.id, "Произошла ошибка парсинга!\nПопробуйте перезапустить парсер.".format(message.from_user, bot.get_me()),
                     parse_mode='html', reply_markup=markup)



def schedule_checker():
    while True:
        schedule.run_pending()
        time.sleep(60)

if __name__ == '__main__':
    print("БОТ ЗАПУЩЕН")
    schedule.every().day.at("18:25").do(parsing_price)
    schedule.every().day.at("18:31").do(parsing_stock)
    Thread(target=schedule_checker).start()
    bot.polling(none_stop=True, interval=0)





print("\nPress any key")
console_admin_input = input()